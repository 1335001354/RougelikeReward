from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def write_and_format_cell(worksheet, cell_address, content, style_dict=None):
    """
    向Excel工作表的指定单元格写入内容并应用格式化
    
    Args:
        worksheet: openpyxl工作表对象
        cell_address: 单元格地址，如 'A1', 'B2' 等
        content: 要写入的内容
        style_dict: 样式字典，包含以下可选键：
            - 'font': 字体样式字典
                - 'name': 字体名称
                - 'size': 字体大小
                - 'bold': 是否加粗 (True/False)
                - 'italic': 是否斜体 (True/False)
                - 'color': 字体颜色 (RGB十六进制字符串，如 'FF0000')
            - 'fill': 填充样式字典
                - 'type': 填充类型 ('solid', 'patternFill' 等)
                - 'color': 背景颜色 (RGB十六进制字符串)
            - 'border': 边框样式字典
                - 'left': 左边框样式 ('thin', 'medium', 'thick' 等)
                - 'right': 右边框样式
                - 'top': 上边框样式
                - 'bottom': 下边框样式
                - 'color': 边框颜色 (RGB十六进制字符串)
            - 'alignment': 对齐样式字典
                - 'horizontal': 水平对齐 ('left', 'center', 'right')
                - 'vertical': 垂直对齐 ('top', 'center', 'bottom')
                - 'wrap_text': 是否自动换行 (True/False)
            - 'number_format': 数字格式字符串
    
    Returns:
        None
    """
    # 写入内容
    worksheet[cell_address] = content
    
    # 如果没有样式信息，直接返回
    if style_dict is None:
        return
    
    # 获取单元格对象
    cell = worksheet[cell_address]
    
    # 应用字体样式
    if 'font' in style_dict:
        font_dict = style_dict['font']
        font = Font()
        
        if 'name' in font_dict:
            font.name = font_dict['name']
        if 'size' in font_dict:
            font.size = font_dict['size']
        if 'bold' in font_dict:
            font.bold = font_dict['bold']
        if 'italic' in font_dict:
            font.italic = font_dict['italic']
        if 'color' in font_dict:
            font.color = font_dict['color']
        
        cell.font = font
    
    # 应用填充样式
    if 'fill' in style_dict:
        fill_dict = style_dict['fill']
        fill = PatternFill()
        
        if 'type' in fill_dict:
            fill.fill_type = fill_dict['type']
        if 'color' in fill_dict:
            fill.fgColor = fill_dict['color']
        
        cell.fill = fill
    
    # 应用边框样式
    if 'border' in style_dict:
        border_dict = style_dict['border']
        border = Border()
        
        # 定义边框样式映射
        border_styles = {
            'thin': Side(style='thin'),
            'medium': Side(style='medium'),
            'thick': Side(style='thick'),
            'dashed': Side(style='dashed'),
            'dotted': Side(style='dotted')
        }
        
        # 应用各边边框
        for side in ['left', 'right', 'top', 'bottom']:
            if side in border_dict:
                style_name = border_dict[side]
                if style_name in border_styles:
                    side_obj = border_styles[style_name]
                    if 'color' in border_dict:
                        side_obj.color = border_dict['color']
                    setattr(border, side, side_obj)
        
        cell.border = border
    
    # 应用对齐样式
    if 'alignment' in style_dict:
        align_dict = style_dict['alignment']
        alignment = Alignment()
        
        if 'horizontal' in align_dict:
            alignment.horizontal = align_dict['horizontal']
        if 'vertical' in align_dict:
            alignment.vertical = align_dict['vertical']
        if 'wrap_text' in align_dict:
            alignment.wrap_text = align_dict['wrap_text']
        
        cell.alignment = alignment
    
    # 应用数字格式
    if 'number_format' in style_dict:
        cell.number_format = style_dict['number_format']

def write_header_cell(worksheet, cell_address, content):
    """
    写入表头单元格的便捷函数
    
    Args:
        worksheet: openpyxl工作表对象
        cell_address: 单元格地址
        content: 要写入的内容
    """
    header_style = {
        'font': {
            'name': '微软雅黑',
            'size': 12,
            'bold': True,
            'color': 'FFFFFF'
        },
        'fill': {
            'type': 'solid',
            'color': '4472C4'
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center'
        },
        'border': {
            'left': 'thin',
            'right': 'thin',
            'top': 'thin',
            'bottom': 'thin',
            'color': '000000'
        }
    }
    
    write_and_format_cell(worksheet, cell_address, content, header_style)

def write_data_cell(worksheet, cell_address, content, is_numeric=False):
    """
    写入数据单元格的便捷函数
    
    Args:
        worksheet: openpyxl工作表对象
        cell_address: 单元格地址
        content: 要写入的内容
        is_numeric: 是否为数值类型
    """
    data_style = {
        'font': {
            'name': '微软雅黑',
            'size': 10
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center'
        },
        'border': {
            'left': 'thin',
            'right': 'thin',
            'top': 'thin',
            'bottom': 'thin'
        }
    }
    
    if is_numeric:
        data_style['number_format'] = '0.0000'
    
    write_and_format_cell(worksheet, cell_address, content, data_style)

def write_percentage_cell(worksheet, cell_address, content):
    """
    写入百分数单元格的便捷函数
    
    Args:
        worksheet: openpyxl工作表对象
        cell_address: 单元格地址
        content: 要写入的内容（小数形式，如0.1234）
    """
    percentage_style = {
        'font': {
            'name': '微软雅黑',
            'size': 10
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center'
        },
        'border': {
            'left': 'thin',
            'right': 'thin',
            'top': 'thin',
            'bottom': 'thin'
        },
        'number_format': '0.00%'
    }
    
    write_and_format_cell(worksheet, cell_address, content, percentage_style)

def write_max_value_cell(worksheet, cell_address, content, gacha_number):
    """
    写入最大值单元格，根据gacha次数设置不同颜色
    
    Args:
        worksheet: openpyxl工作表对象
        cell_address: 单元格地址
        content: 要写入的内容
        gacha_number: gacha次数（12, 15, 或 20）
    """
    # 根据gacha次数设置不同的背景颜色
    color_map = {
        12: 'FFB6C1',  # 浅粉色
        15: '98FB98',  # 浅绿色
        20: '87CEEB'   # 浅蓝色
    }
    
    max_value_style = {
        'font': {
            'name': '微软雅黑',
            'size': 10,
            'bold': True
        },
        'fill': {
            'type': 'solid',
            'color': color_map.get(gacha_number, 'FFFFFF')
        },
        'alignment': {
            'horizontal': 'center',
            'vertical': 'center'
        },
        'border': {
            'left': 'thin',
            'right': 'thin',
            'top': 'thin',
            'bottom': 'thin'
        },
        'number_format': '0.00%'
    }
    
    write_and_format_cell(worksheet, cell_address, content, max_value_style)

def write_gacha_legend(worksheet, start_row, start_col):
    """
    在表格下方写入gacha次数图例
    
    Args:
        worksheet: openpyxl工作表对象
        start_row: 开始行号
        start_col: 开始列号
    """
    gacha_info = [
        (12, 'FFB6C1', '浅粉色'),
        (15, '98FB98', '浅绿色'),
        (20, '87CEEB', '浅蓝色')
    ]
    
    for i, (gacha_num, color, color_name) in enumerate(gacha_info):
        row = start_row + i
        
        # 写入颜色说明
        legend_style = {
            'font': {
                'name': '微软雅黑',
                'size': 10
            },
            'fill': {
                'type': 'solid',
                'color': color
            },
            'alignment': {
                'horizontal': 'center',
                'vertical': 'center'
            },
            'border': {
                'left': 'thin',
                'right': 'thin',
                'top': 'thin',
                'bottom': 'thin'
            }
        }
        
        write_and_format_cell(worksheet, f'{get_column_letter(start_col)}{row}', 
                             f'Gacha {gacha_num}', legend_style)
        
        # 写入说明文字
        text_style = {
            'font': {
                'name': '微软雅黑',
                'size': 10
            },
            'alignment': {
                'horizontal': 'left',
                'vertical': 'center'
            }
        }
        
        write_and_format_cell(worksheet, f'{get_column_letter(start_col + 1)}{row}', 
                             f'最大值标记 ({color_name})', text_style)

def apply_data_bars(worksheet, start_row, start_col, end_row, end_col):
    """
    为指定区域添加蓝色数据条效果（通过条件格式化实现）
    
    Args:
        worksheet: openpyxl工作表对象
        start_row: 开始行号
        start_col: 开始列号
        end_row: 结束行号
        end_col: 结束列号
    """
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Color
    
    # 创建数据条规则
    data_bar_rule = DataBarRule(
        start_type="min",
        end_type="max",
        color=Color(rgb="4472C4"),  # 蓝色
        showValue=True,
        minLength=0,
        maxLength=100
    )
    
    # 应用条件格式化
    cell_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    worksheet.conditional_formatting.add(cell_range, data_bar_rule)
