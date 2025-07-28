from character import Character
from ProDistribution import ProDistribution
from type import Style
import random
import pandas as pd
from write import write_header_cell, write_percentage_cell, write_max_value_cell, write_gacha_legend, apply_data_bars

"""
游戏模拟器系统

功能说明：
- 模拟角色抽卡过程
- 支持多轮统计和Excel文件生成
- 提供概率分布计算和数据分析
"""

class Simulator:
    """游戏模拟器"""
    
    def __init__(self):
        """初始化模拟器"""
        self.pro_distribution = ProDistribution()
        self.characters = []
        self.initialize_characters()
    
    def initialize_characters(self):
        """初始化六个角色"""
        # 创建六个角色：两个5级，两个18级，两个35级
        # 每种等级分别设置havetool为true和false，属性均为熔岩球
        
        # 5级角色
        char_5_false = Character("熔岩球", 5, havetool=False)
        char_5_true = Character("熔岩球", 5, havetool=True)
        
        # 18级角色
        char_18_false = Character("熔岩球", 18, havetool=False)
        char_18_true = Character("熔岩球", 18, havetool=True)
        
        # 35级角色
        char_35_false = Character("熔岩球", 35, havetool=False)
        char_35_true = Character("熔岩球", 35, havetool=True)
        
        # 添加到角色列表
        self.characters = [
            char_5_false, char_5_true,
            char_18_false, char_18_true,
            char_35_false, char_35_true
        ]
    
    def _perform_single_draw(self, character, enable_reroll=True, remaining_rerolls=0):
        """
        执行单次抽卡，包含重新roll功能
        
        Args:
            character: 角色对象
            enable_reroll: 是否启用重新roll功能，默认True
            remaining_rerolls: 剩余重新roll次数，默认0次
            
        Returns:
            tuple: (selected_style, three_cards, reroll_count, reroll_history, remaining_rerolls)
        """
        reroll_count = 0
        reroll_history = []
        
        while True:
            # 获取当前概率分布
            probabilities = self.pro_distribution.get_current_weight(character)
            styles = list(probabilities.keys())
            probs = list(probabilities.values())
            
            # 使用random.choices进行加权随机选择三个流派
            three_cards = random.choices(styles, weights=probs, k=3)
            
            # 角色选择逻辑：优先选择和自身属性相同的流派
            character_attribute = character.attribute
            selected_style = None
            
            # 检查是否有与自身属性相同的流派
            for card in three_cards:
                if card == character_attribute:
                    selected_style = card
                    break
            
            # 如果没有找到相同属性的流派，选择第一个
            if selected_style is None:
                selected_style = three_cards[0]
            
            # 检查是否需要重新roll
            should_reroll = False
            if enable_reroll and remaining_rerolls > 0:
                # 条件1：三张卡牌全不是自身流派
                no_main_style = character_attribute not in three_cards
                
                # 条件2：自身已拿到主攻流派或目前还没凑齐三个流派
                has_main_style = character.get_attribute_value(character_attribute) > 0
                style_count_less_than_3 = character.get_style_count() < 3
                
                should_reroll = no_main_style and (has_main_style or style_count_less_than_3)
            
            if should_reroll:
                reroll_count += 1
                remaining_rerolls -= 1
                reroll_history.append(three_cards.copy())
                continue
            else:
                break
        
        return selected_style, three_cards, reroll_count, reroll_history, remaining_rerolls

    def simulate_card_drawing(self, draw_count=15, enable_reroll=True, max_rerolls=2):
        """
        模拟抽卡过程
        
        Args:
            draw_count: 抽卡次数，默认15次
            enable_reroll: 是否启用重新roll功能，默认True
            max_rerolls: 一轮完整模拟中最大重新roll次数，默认2次
        """
        print(f"\n=== 模拟{draw_count}次抽卡 ===")
        if enable_reroll:
            print(f"重新roll功能: 启用 (一轮最多{max_rerolls}次)")
        else:
            print("重新roll功能: 禁用")
        print("=" * 80)
        
        for i, character in enumerate(self.characters, 1):
            print(f"\n角色 {i} (等级{character.get_level()}, havetool={character.get_havetool()}):")
            
            # 记录初始状态
            initial_style_count = character.get_style_count()
            print(f"  初始流派拥有数量: {initial_style_count}")
            
            # 模拟抽卡过程
            draw_results = []
            remaining_rerolls = max_rerolls  # 初始化剩余重新roll次数
            
            for draw in range(draw_count):
                # 获取当前概率分布
                probabilities = self.pro_distribution.get_current_weight(character)
                # 展示当前各个流派的权重
                prob_str = ', '.join([f"{k}:{v:.3f}" for k, v in probabilities.items()])
                print(f"    第{draw+1}次抽卡前流派权重: {prob_str}")
                if enable_reroll:
                    print(f"    剩余重新roll次数: {remaining_rerolls}")
                
                # 执行单次抽卡
                selected_style, three_cards, reroll_count, reroll_history, remaining_rerolls = self._perform_single_draw(
                    character, enable_reroll, remaining_rerolls
                )
                
                # 记录抽卡结果
                draw_result = {
                    'cards': three_cards,
                    'selected': selected_style,
                    'reroll_count': reroll_count,
                    'reroll_history': reroll_history
                }
                draw_results.append(draw_result)
                
                # 增加对应流派的value
                character.increase_attribute_value(selected_style, 1)
                
                # 每5次抽卡显示一次进度
                if (draw + 1) % 5 == 0:
                    current_style_count = character.get_style_count()
                    print(f"    第{draw + 1}次抽卡后: 流派拥有数量 = {current_style_count}")
            
            # 显示最终结果
            final_style_count = character.get_style_count()
            print(f"  最终流派拥有数量: {final_style_count}")
            print(f"  流派获得情况:")
            
            # 显示所有流派的最终值
            for style, value in character.attribute_values.items():
                if value > 0:
                    print(f"    {style}: {value}")
            
            # 显示抽卡历史
            print(f"  抽卡历史:")
            for j, result in enumerate(draw_results, 1):
                cards_str = ", ".join(result['cards'])
                selected = result['selected']
                selection_reason = "优先选择" if selected == character.attribute else "默认选择"
                reroll_info = f" (重新roll{result['reroll_count']}次)" if result['reroll_count'] > 0 else ""
                print(f"    第{j:2d}次: [{cards_str}] -> {selected} ({selection_reason}){reroll_info}")
            
            print(f"  流派数量变化: {initial_style_count} -> {final_style_count}")
    
    def simulate_attribute_value_ratio(self, draw_count=15, rounds=10, threshold=7, enable_reroll=True, max_rerolls=2):
        """
        模拟多轮并统计角色自身流派属性值大于threshold的比例
        
        Args:
            draw_count: 每轮抽卡次数，默认15次
            rounds: 模拟轮数，默认10轮
            threshold: 阈值，默认7
            enable_reroll: 是否启用重新roll功能，默认True
            max_rerolls: 一轮完整模拟中最大重新roll次数，默认2次
        """
        print(f"\n=== 模拟{rounds}轮，每轮{draw_count}次抽卡，统计自身流派属性值>{threshold}的比例 ===")
        if enable_reroll:
            print(f"重新roll功能: 启用 (一轮最多{max_rerolls}次)")
        else:
            print("重新roll功能: 禁用")
        print("=" * 80)
        for i, character in enumerate(self.characters, 1):
            success_count = 0
            no_main_attribute_count = 0  # 统计没有主攻属性的轮数
            print(f"\n角色 {i} (等级{character.get_level()}, havetool={character.get_havetool()}):")
            for round_idx in range(1, rounds+1):
                character.reset_all_attributes()
                remaining_rerolls = max_rerolls  # 初始化剩余重新roll次数
                for _ in range(draw_count):
                    # 执行单次抽卡
                    selected_style, _, _, _, remaining_rerolls = self._perform_single_draw(
                        character, enable_reroll, remaining_rerolls
                    )
                    character.increase_attribute_value(selected_style, 1)
                
                # 检查本轮结束时是否有主攻属性
                if character.get_attribute_value(character.attribute) == 0:
                    no_main_attribute_count += 1
                
                if character.get_attribute_value(character.attribute) > threshold:
                    success_count += 1
            
            ratio = success_count / rounds
            no_main_ratio = no_main_attribute_count / rounds
            print(f"  自身流派属性值>{threshold}的比例 = {ratio:.3f}")
            print(f"  属性池没有主攻属性的比例 = {no_main_ratio:.3f}")

    def simulate_and_generate_excel(self, draw_count=15, rounds=1000, enable_reroll=True, max_rerolls=2):
        """
        模拟多轮抽卡并生成Excel文件
        
        Args:
            draw_count: 每轮抽卡次数，默认15次
            rounds: 模拟轮数，默认1000轮
            enable_reroll: 是否启用重新roll功能，默认True
            max_rerolls: 一轮完整模拟中最大重新roll次数，默认2次
            
        Returns:
            None，生成Excel文件到当前目录
        """
        print(f"\n=== 模拟{rounds}轮，每轮{draw_count}次抽卡，生成Excel文件 ===")
        if enable_reroll:
            print(f"重新roll功能: 启用 (一轮最多{max_rerolls}次)")
        else:
            print("重新roll功能: 禁用")
        print("=" * 80)
        
        # 初始化统计数据
        character_stats = []
        for i in range(len(self.characters)):
            character_stats.append({})
            for j in range(draw_count + 1):
                character_stats[i][j] = {}
        
        # 执行多轮模拟
        for round_idx in range(rounds):
            # 重置所有角色属性
            for character in self.characters:
                character.reset_all_attributes()
            
            # 初始化每个角色的剩余重新roll次数
            remaining_rerolls_per_character = [max_rerolls] * len(self.characters)
            
            # 执行抽卡过程
            for draw in range(draw_count):
                for i, character in enumerate(self.characters):
                    # 执行单次抽卡
                    selected_style, _, _, _, remaining_rerolls_per_character[i] = self._perform_single_draw(
                        character, enable_reroll, remaining_rerolls_per_character[i]
                    )
                    
                    # 增加对应流派的value
                    character.increase_attribute_value(selected_style, 1)
                    
                    # 记录当前状态
                    current_value = character.get_attribute_value(character.attribute)
                    if current_value not in character_stats[i][draw + 1]:
                        character_stats[i][draw + 1][current_value] = 0
                    character_stats[i][draw + 1][current_value] += 1
        
        # 生成Excel文件
        self._generate_excel_files(character_stats, draw_count, rounds)
        
        print("Excel文件生成完成！")
    
    def _generate_excel_files(self, character_stats, draw_count, rounds):
        """
        生成Excel文件，包含所有角色的抽卡统计数据
        
        Args:
            character_stats: 角色统计数据字典
            draw_count: 抽卡次数
            rounds: 模拟轮数
            
        Returns:
            None，生成Excel文件到当前目录
        """
        filename = "simulation_results.xlsx"
        print(f"  生成文件: {filename}")
        
        # 按等级分组角色
        level_groups = {}
        for i, character in enumerate(self.characters):
            level = character.get_level()
            if level not in level_groups:
                level_groups[level] = []
            level_groups[level].append((i, character))
        
        # 创建Excel写入器
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 按指定顺序处理等级：35, 18, 5
            level_order = [35, 18, 5]
            for level in level_order:
                if level not in level_groups:
                    continue
                characters = level_groups[level]
                # 创建工作表名称
                sheet_name = f"等级{level}"
                
                # 获取该等级所有角色可能的属性值
                all_values = set()
                for char_idx, character in characters:
                    for draw in range(draw_count + 1):
                        all_values.update(character_stats[char_idx][draw].keys())
                all_values = sorted(list(all_values))
                
                # 创建列名（包含两种tool状态）
                columns = [''] + [f"The {j}th gacha" for j in range(draw_count + 1)]
                
                # 为每个tool状态创建数据
                tool_data = {}
                for char_idx, character in characters:
                    havetool_text = "有灵器" if character.get_havetool() else "无灵器"
                    
                    data = []
                    for value in all_values:
                        row = [f"aimming_level{value}"]
                        for draw in range(draw_count + 1):
                            count = character_stats[char_idx][draw].get(value, 0)
                            ratio = count / rounds
                            row.append(ratio)
                        data.append(row)
                    
                    tool_data[havetool_text] = data
                
                # 创建上下摆放的DataFrame
                merged_data = []
                merged_columns = [''] + [f"The {j}th gacha" for j in range(draw_count + 1)]
                
                # 添加无灵器数据
                if "无灵器" in tool_data:
                    # 添加无灵器标题行
                    title_row = ["无灵器"] + [""] * draw_count
                    merged_data.append(title_row)
                    
                    # 添加无灵器数据
                    for value in all_values:
                        row = [f"aimming_level{value}"]
                        row.extend(tool_data["无灵器"][all_values.index(value)][1:])
                        merged_data.append(row)
                    
                    # 添加空行分隔
                    if "有灵器" in tool_data:
                        empty_row = [""] * (draw_count + 1)
                        merged_data.append(empty_row)
                
                # 添加有灵器数据
                if "有灵器" in tool_data:
                    # 添加有灵器标题行
                    title_row = ["有灵器"] + [""] * draw_count
                    merged_data.append(title_row)
                    
                    # 添加有灵器数据
                    for value in all_values:
                        row = [f"aimming_level{value}"]
                        row.extend(tool_data["有灵器"][all_values.index(value)][1:])
                        merged_data.append(row)
                
                # 创建DataFrame
                df = pd.DataFrame(merged_data, columns=merged_columns)
                
                # 写入Excel工作表
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 获取工作表对象进行格式化
                worksheet = writer.sheets[sheet_name]
                
                # 插入角色信息行
                worksheet.insert_rows(1, 3)
                
                # 写入角色信息
                from openpyxl.utils import get_column_letter
                
                # 角色信息样式
                info_style = {
                    'font': {
                        'name': '微软雅黑',
                        'size': 12,
                        'bold': True
                    },
                    'fill': {
                        'type': 'solid',
                        'color': 'E6E6FA'
                    }
                }
                
                from write import write_and_format_cell
                write_and_format_cell(worksheet, 'A1', f"等级 {level} 角色数据", info_style)
                write_and_format_cell(worksheet, 'A2', f"包含无灵器和有灵器两种状态", info_style)
                
                # 写入表头
                for col_idx, col_name in enumerate(merged_columns, 1):
                    cell_address = f"{get_column_letter(col_idx)}4"
                    write_header_cell(worksheet, cell_address, col_name)
                
                # 写入数据并格式化
                data_start_row = 5
                data_start_col = 2  # 从B列开始（跳过第一列标签）
                
                # 找到gacha=12, 15, 20对应的列索引
                gacha_columns = {}
                for col_idx, col_name in enumerate(merged_columns):
                    if any(gacha in col_name for gacha in ["12th", "15th", "20th"]):
                        # 从列名中提取gacha数字
                        for gacha_num in [12, 15, 20]:
                            if f"{gacha_num}th" in col_name:
                                gacha_columns[gacha_num] = col_idx
                                break
                
                # 分别收集无灵器和有灵器在gacha=12, 15, 20列的最大值
                max_values = {}
                
                # 直接从tool_data中获取数据，避免索引问题
                for gacha_num in [12, 15, 20]:
                    if gacha_num in gacha_columns:
                        col_idx = gacha_columns[gacha_num]
                        max_values[gacha_num] = {'no_tool': None, 'has_tool': None}
                        
                        # 找无灵器最大值
                        if "无灵器" in tool_data:
                            no_tool_max_value = 0
                            no_tool_max_cells = []
                            for row_idx, row_data in enumerate(tool_data["无灵器"]):
                                ratio = row_data[col_idx]
                                # 确保ratio是数值类型
                                if isinstance(ratio, str):
                                    try:
                                        ratio = float(ratio)
                                    except ValueError:
                                        continue
                                if ratio > no_tool_max_value:
                                    no_tool_max_value = ratio
                                    # 计算实际的行号（跳过标题行）
                                    actual_row = data_start_row + 1 + row_idx
                                    no_tool_max_cells = [f"{get_column_letter(col_idx + 1)}{actual_row}"]
                                elif ratio == no_tool_max_value:
                                    actual_row = data_start_row + 1 + row_idx
                                    no_tool_max_cells.append(f"{get_column_letter(col_idx + 1)}{actual_row}")
                            
                            max_values[gacha_num]['no_tool'] = {
                                'value': no_tool_max_value,
                                'cells': no_tool_max_cells
                            }
                        
                        # 找有灵器最大值
                        if "有灵器" in tool_data:
                            has_tool_max_value = 0
                            has_tool_max_cells = []
                            for row_idx, row_data in enumerate(tool_data["有灵器"]):
                                ratio = row_data[col_idx]
                                # 确保ratio是数值类型
                                if isinstance(ratio, str):
                                    try:
                                        ratio = float(ratio)
                                    except ValueError:
                                        continue
                                if ratio > has_tool_max_value:
                                    has_tool_max_value = ratio
                                    # 计算实际的行号（跳过标题行、数据行、空行、标题行）
                                    actual_row = data_start_row + 1 + len(all_values) + 2 + row_idx
                                    has_tool_max_cells = [f"{get_column_letter(col_idx + 1)}{actual_row}"]
                                elif ratio == has_tool_max_value:
                                    actual_row = data_start_row + 1 + len(all_values) + 2 + row_idx
                                    has_tool_max_cells.append(f"{get_column_letter(col_idx + 1)}{actual_row}")
                            
                            max_values[gacha_num]['has_tool'] = {
                                'value': has_tool_max_value,
                                'cells': has_tool_max_cells
                            }
                
                # 写入数据行
                current_row = data_start_row
                
                # 处理无灵器数据
                if "无灵器" in tool_data:
                    # 写入无灵器标题
                    title_cell = f"A{current_row}"
                    title_style = {
                        'font': {
                            'name': '微软雅黑',
                            'size': 12,
                            'bold': True
                        },
                        'fill': {
                            'type': 'solid',
                            'color': 'FFE6CC'
                        }
                    }
                    write_and_format_cell(worksheet, title_cell, "无灵器", title_style)
                    current_row += 1
                    
                    # 写入无灵器数据
                    for row_idx, value in enumerate(all_values):
                        # 写入行标签
                        label_cell = f"A{current_row}"
                        write_and_format_cell(worksheet, label_cell, f"aimming_level{value}")
                        
                        for col_idx, col_name in enumerate(merged_columns[1:], 1):  # 跳过第一列
                            cell_address = f"{get_column_letter(col_idx + 1)}{current_row}"
                            ratio = merged_data[current_row - data_start_row][col_idx]
                            
                            # 检查是否为gacha=12, 15, 20的列
                            gacha_num = None
                            for gacha_col in gacha_columns:
                                if col_idx == gacha_columns[gacha_col]:
                                    gacha_num = gacha_col
                                    break
                            
                            # 写入数据
                            if gacha_num is not None:
                                # 先写入普通格式，稍后更新为最大值格式
                                write_percentage_cell(worksheet, cell_address, ratio)
                            else:
                                write_percentage_cell(worksheet, cell_address, ratio)
                        
                        current_row += 1
                    
                    # 添加空行分隔
                    if "有灵器" in tool_data:
                        current_row += 1
                
                # 处理有灵器数据
                if "有灵器" in tool_data:
                    # 写入有灵器标题
                    title_cell = f"A{current_row}"
                    title_style = {
                        'font': {
                            'name': '微软雅黑',
                            'size': 12,
                            'bold': True
                        },
                        'fill': {
                            'type': 'solid',
                            'color': 'E6F3FF'
                        }
                    }
                    write_and_format_cell(worksheet, title_cell, "有灵器", title_style)
                    current_row += 1
                    
                    # 写入有灵器数据
                    for row_idx, value in enumerate(all_values):
                        # 写入行标签
                        label_cell = f"A{current_row}"
                        write_and_format_cell(worksheet, label_cell, f"aimming_level{value}")
                        
                        for col_idx, col_name in enumerate(merged_columns[1:], 1):  # 跳过第一列
                            cell_address = f"{get_column_letter(col_idx + 1)}{current_row}"
                            ratio = merged_data[current_row - data_start_row][col_idx]
                            
                            # 检查是否为gacha=12, 15, 20的列
                            gacha_num = None
                            for gacha_col in gacha_columns:
                                if col_idx == gacha_columns[gacha_col]:
                                    gacha_num = gacha_col
                                    break
                            
                            # 写入数据
                            if gacha_num is not None:
                                # 先写入普通格式，稍后更新为最大值格式
                                write_percentage_cell(worksheet, cell_address, ratio)
                            else:
                                write_percentage_cell(worksheet, cell_address, ratio)
                        
                        current_row += 1
                
                # 重新写入最大值单元格并设置颜色
                for gacha_num, max_info in max_values.items():
                    # 设置无灵器最大值
                    if max_info['no_tool']:
                        for cell_address in max_info['no_tool']['cells']:
                            write_max_value_cell(worksheet, cell_address, max_info['no_tool']['value'], gacha_num)
                    
                    # 设置有灵器最大值
                    if max_info['has_tool']:
                        for cell_address in max_info['has_tool']['cells']:
                            write_max_value_cell(worksheet, cell_address, max_info['has_tool']['value'], gacha_num)
                
                # 添加数据条
                data_end_row = current_row - 1
                data_end_col = len(merged_columns)
                apply_data_bars(worksheet, data_start_row, data_start_col, data_end_row, data_end_col)
                
                # 添加图例
                legend_start_row = data_end_row + 3
                write_gacha_legend(worksheet, legend_start_row, 1)



def main():
    """主函数"""
    simulator = Simulator()
    
    # 示例1：启用重新roll功能（默认设置）
    print("=== 示例1：启用重新roll功能（一轮最多2次） ===")
    simulator.simulate_card_drawing(draw_count=10, enable_reroll=True, max_rerolls=2)
    
    # 示例2：禁用重新roll功能
    print("\n" + "="*80)
    print("=== 示例2：禁用重新roll功能 ===")
    simulator.simulate_card_drawing(draw_count=10, enable_reroll=False)
    
    # 示例3：自定义重新roll次数
    print("\n" + "="*80)
    print("=== 示例3：自定义重新roll次数（一轮最多3次） ===")
    simulator.simulate_card_drawing(draw_count=10, enable_reroll=True, max_rerolls=3)
    
    
    # 示例4：生成Excel文件（启用重新roll）
    print("\n" + "="*80)
    print("=== 示例4：生成Excel文件（启用重新roll，一轮最多2次） ===")
    simulator.simulate_and_generate_excel(draw_count=20, rounds=1000, enable_reroll=True, max_rerolls=2)
    

if __name__ == "__main__":
    main()
